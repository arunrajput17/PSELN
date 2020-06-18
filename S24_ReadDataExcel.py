### Reading Data from MS-Excel using OpenPyXL 

import openpyxl

path= 'D:\\Day Shift\\Zpyth\\Pseln\\TestFiles\\data24.xlsx'

workbook = openpyxl.load_workbook(path) # creating workbook object

### Getting sheet from workbook method 1 (if only one sheet in excel)
# sheet = workbook.active

### Getting sheet from workbook Method 2 ( multiple sheets in excel)
sheet = workbook.get_sheet_by_name('Sheet1')


rows=sheet.max_row      # Counting Rows
columns = sheet.max_column      # counting columns

print(rows)
print(columns)

# Reading data from sheet

for r in range (1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(row=r,column=c).value, end='        ')
    print()




